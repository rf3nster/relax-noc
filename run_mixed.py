############################################################################
# Script:
#    Runs RELAX testbench in mixed mode, places results in CSVs and
#    .xlsx files
#
# Requires: VHDL-2008, NON-SYNTHESIZABLE
# 
# Copyright 2021 Rick Fenster
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
# 
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
############################################################################

# Imports
import subprocess
import random
import sys
import os
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
print("================================================================================")  
print("**** Preparing Environment ****")  
# Grab injection rates from file
try:
    injrates_file = open ('injectionrates.dat','r')
except OSError:
    print ("... Cannot read injection rates")
    quit()
injrates = injrates_file.readlines()
# Remove trailing linebreaks and create list
injrates = [s.replace("\n","") for s in injrates]
del injrates_file

# Get weights to test for
try:
    weights_file = open ('weights.dat','r')
except OSError:
    print ("... Cannot read data type weights")
    quit()    
weights = weights_file.readlines()
# Remove trailing linebreaks and create list
weights = [s.replace("\n","") for s in weights]
del weights_file
weights = [s.split(",") for s in weights]
# Parameters
params_file = open('./src/shared/noc_parameterspkg.vhd')
# Get X Size of Network
x_size = params_file.readlines()[43]
x_size = re.sub("[^0-9]", "", x_size)
# Get Y Size of Network
params_file = open('./src/shared/noc_parameterspkg.vhd')
y_size = params_file.readlines()[46]
y_size = re.sub("[^0-9]", "", y_size)
# Get quantity of packets
params_file = open('./src/shared/noc_parameterspkg.vhd')
packet_qty = params_file.readlines()[37]
packet_qty = re.sub("[^0-9]", "", packet_qty)
# Get number of ticks per period
params_file = open('./src/shared/noc_parameterspkg.vhd')
period_size = params_file.readlines()[40]
period_size = re.sub("[^0-9]", "", period_size)
# Clean up
del params_file
 
try:
    os.mkdir("csv_data")    
except OSError:
    print ("... Creation of the compiled csv data directory failed or already exists")
try:
    os.mkdir("transcript_data")    
except OSError:
    print ("... Creation of the transcript data directory failed or already exists")
# Clean simulation env
print("... Cleaning Environment")
# Clean transcript CSVs
dir_name = os.getcwd()+ "/transcript_data"
current_csv_files = os.listdir(dir_name)
for item in current_csv_files:
    if item.endswith(".csv"):
        os.remove(os.path.join(dir_name, item))
# Clean compiled CSVs
dir_name = os.getcwd()+ "/csv_data"
current_csv_files = os.listdir(dir_name)
for item in current_csv_files:
    if item.endswith(".csv"):
        os.remove(os.path.join(dir_name, item))
dir_name = os.getcwd()
current_csv_files = os.listdir(dir_name) 
for item in current_csv_files:
    if item.endswith(".csv"):
        os.remove(os.path.join(dir_name, item))
del current_csv_files
# Clean XLSXs
current_csv_files = os.listdir(dir_name) 
for item in current_csv_files:
    if item.endswith(".xlsx"):
        os.remove(os.path.join(dir_name, item))
del current_csv_files
# Clean WLFs
current_wlf_files = os.listdir(dir_name)
for item in current_wlf_files:
    if item.endswith(".wlf"):
        os.remove(os.path.join(dir_name, item))
# Print the test conditions
print("================================================================================")   
print("**** Test Parameters ****")
print("Test type: Weighted Random")
print("X Size of Network:", x_size)
print("Y Size of Network:", y_size)
print("Number of Packets to Inject:",packet_qty)
print("Number of Ticks Per Period:",period_size)
print("Number of injection rates to test:", len(injrates))
print("Number of weighted scenarios:", len(weights))
input("Press any key to continue...")


# First loop, based on the weights
for weight_cnt in range(len(weights)):
    print("================================================================================")   
    print("Accurate Weight: "+str(weights[weight_cnt][0])+"%, Approximate Weight: "+str(weights[weight_cnt][1]+"%"))
   
    # Write weight values to parameters file
    params_file = open('./src/shared/noc_parameterspkg.vhd')
    params_data = params_file.readlines()
    params_data[60] = '    constant acc_data_weight : natural := '+str(weights[weight_cnt][0])+ ";\n"
    params_data[61] = '    constant apx_data_weight : natural := '+str(weights[weight_cnt][1])+ ";\n"
    params_file = open('./src/shared/noc_parameterspkg.vhd','w')
    params_file.writelines(params_data)
    params_file.close() 

    # Start passes of differnet injection rates
    for injrate_cnt in range(len(injrates)):

        print("**** Current injection rate is: "+str(injrates[injrate_cnt]+" ****"))
        # Update parameters file to reflect injection rate
        current_injection_rate = int((float(injrates[injrate_cnt]) * int(period_size)))

        # Write weight values to parameters file
        params_file = open('./src/shared/noc_parameterspkg.vhd')
        params_data = params_file.readlines()
        params_data[82] = '    constant packets_per_period : integer := '+str(current_injection_rate)+ ";\n"
        params_file = open('./src/shared/noc_parameterspkg.vhd','w')
        params_file.writelines(params_data)
        params_file.close()         
        
        # Add random seeds
        rand_seed_ext = random.random()
        pe_inject_file = open("./src/pe/pe_injectiondriver.vhd", 'r')
        pe_inject_file_lines = pe_inject_file.readlines()
        pe_inject_file.close()
        pe_inject_file_lines[138] = "                   injectionTimes_rand.InitSeed(injectionTimes_rand'instance_name & to_string(x_coord)&to_string(y_coord) &\""+str(rand_seed_ext)+"\");\n"
        rand_seed_ext = random.random()
        pe_inject_file_lines[158] = "                   randType.InitSeed(randDest'instance_name&to_string(x_coord)&to_string(y_coord) &\""+str(rand_seed_ext)+"\");\n"
        pe_inject_file = open("./src/pe/pe_injectiondriver.vhd", 'w')
        pe_inject_file.writelines(pe_inject_file_lines)
        pe_inject_file.close()
        del pe_inject_file_lines
        ############################################################################################################
        # Update these lines with your simulator command
        # Launch VSIM
        print("... Running Modelsim")
        subprocess.call( ["vsim","-c",  "-do", "do ./test_scripts/noc_mixed_weighted.do"], cwd=os.getcwd(), stdout=subprocess.DEVNULL,  stderr=subprocess.DEVNULL)
        ############################################################################################################
        print("... Concatenating results")
        if (injrates[injrate_cnt]) == 1.00 :
            input("At Inj 1.00, press enter to continue")
        # Concatenate accurate results if acc != 0
        results_write_file_name = "results_"+str(injrates[injrate_cnt])+".csv"
        results_write_file = open('./csv_data/'+results_write_file_name,'a')
        for x_coord in range (int(x_size)):
            for y_coord in range (int(y_size)):
                result_read_file_name = "./transcript_data/transcript_received_"+str(x_coord)+"_"+str(y_coord)+".csv"
                results_read_file = open(result_read_file_name)
                results_write_file.writelines(results_read_file.readlines())

        # Clean up objects that are no longer used
        del results_write_file_name
        del results_write_file
        del result_read_file_name
        del results_read_file
        # Delete remaining CSVs
        dir_name = os.getcwd()
        # Clean CSVs
        current_csv_files = os.listdir(dir_name)
        for item in current_csv_files:
            if item.endswith(".csv"):
                os.remove(os.path.join(dir_name, item))
        del current_csv_files
        # Clean WLFs
        current_wlf_files = os.listdir(dir_name)
        for item in current_wlf_files:
            if item.endswith(".wlf"):
                os.remove(os.path.join(dir_name, item))
    
    # Init lists for averages
    averages = []
    print("**** Creating Excel Workbook ****")
    # Create workbook
    print("... Adding Data to Book")
    wb = Workbook(write_only=True)

    for injrates_cnt in range(len(injrates)):
        
        ws1 = wb.create_sheet("IR = "+str(injrates[injrates_cnt]))
        # Add header
        ws1.append(["ID","Type","Packet Sent At", "Packet Received At", "Received X Co-Ord", "Received Y Co-Ord", "Delta CC"])

        # Grab CSV file
        csv_results_file = open('./csv_data/results_'+(injrates[injrates_cnt].strip())+'.csv', 'r')
        # Traverse the CSV file
        line = csv_results_file.readline()
        cnt = 0

        while line:

            # Strip string of newline
            line = line.split(',')
            line = [i.replace('\n','') for i in line]
            # Convert values to int
            line[0] = int(line[0])
            line[2] = int(line[2])
            line[3] = int(line[3])
            line[4] = int(line[4])
            line[5] = int(line[5])
            # Add subtraction
            line.append("=D"+str(cnt+2)+"-C"+str(cnt+2))
            # Append to sheet
            ws1.append(line)

            # Go to new line
            line = csv_results_file.readline()
            cnt += 1

        # Add Average
        ws1.append(["Average","","","","","","=AVERAGE(G2:G"+str(cnt+1)+")"])
        # Append average to list
        averages.append("G"+str(cnt+2))

    # Create Sheet with Results
    print("... Creating Results Sheet")
    ws1 = wb.create_sheet("Results and System Info")
    # System Info
    ws1.append(["System Info"])
    ws1.append(["X Size", x_size])
    ws1.append(["Y Size", y_size])
    ws1.append(["Number of Injected Packets Per Node", packet_qty])
    ws1.append(["Number of Ticks Per Period",period_size])
    ws1.append(["Accurate Weight",weights[weight_cnt][0]])
    ws1.append(["Approximate Weight",weights[weight_cnt][1]])
    # Plot Results
    ws1.append(["Results"])
    ws1.append(["Injection Rate", "Latency"])
    # Add Averages to sheet
    for avg_cnt in range(len(averages)):
            ws1.append([str(injrates[avg_cnt]), "='IR = "+str(injrates[avg_cnt])+"'!"+averages[avg_cnt]])
    wb.save (filename = "results_mixed_weighted_acc"+str(weights[weight_cnt][0])+"_apx"+str(weights[weight_cnt][1])+".xlsx")  
